import pandas as pd
import openpyxl as xl
from shutil import copy
from openpyxl.styles import Alignment
from isdayoff import ProdCalendar, DayType, date as datee
from datetime import date, timedelta
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import datetime
from openpyxl.styles import PatternFill



def sutuation_obosration(night_days_second_half,count_of_work_days,day_days_first_half,night_days_first_half,count_of_work_hours,day_hours_first_half,night_hours_first_half,day_hours_second_half,night_hours_second_half,day_days_second_half,r):
        for i in sheet[f'E{r}:T{r}'][0]:   #берем строчку, где написано, был/не был на работе,  и т.п.
            Workday = i.value                                                           #берем  первый элемент и проверяем
            if(Workday != 'ДР' and Workday != 'Я' and Workday != 'В' and  Workday != None and Workday != "О"  and Workday != "Б"):
                day_days_first_half+=1
                night_days_first_half+=1      
            elif Workday == 'ДР':
                count_of_work_days += 1
            elif Workday == 'Я':
                count_of_work_days += 1
        return count_of_work_days
     
        # добавляем в отведенные ячейки сотрудника кол-во отработанных дней

        if (file != "Отдел информационных технологий.xlsx"):
            sheet[cell_function_days_U(cell_num_days)] = count_of_work_days 
            
        else:
            sheet[cell_function_days_U(cell_num_days)] = str("{:.0f}".format(count_of_work_days + day_days_first_half)) + "/" + str("{:.0f}".format(night_days_first_half))
            
        for i in sheet[f'E{r+1}:T{r+1}'][0]:    
            if(i.value !=None): #в месяце могут быть не все 31, а значит могут быть пустые клетки, тобиш None, их складывать нельзя
                Workday = i.value
                if(isinstance(Workday,int)==True or isinstance(Workday,float)==True):
                    count_of_work_hours += Workday                                    
                else:
                    night_hours_first_half+=4 * day_fraction_for_lazy_ass
                    day_hours_first_half += 2 * day_fraction_for_lazy_ass
                    
        if (file != "Отдел информационных технологий.xlsx"):
            sheet[cell_function_hours_U(cell_num_hours)] =  count_of_work_hours
        else:   
            sheet[cell_function_hours_U(cell_num_hours)] =  str("{:.0f}".format(count_of_work_hours+day_hours_first_half)) + "/" + str("{:.0f}".format(night_hours_first_half))
        
        count_of_work_hours_half = count_of_work_hours #считаем кол-во рабочих часов для первых 15 дней и "запоминаем их"
        count_of_work_hours = 0 #обнуляем счетчик, чтобы посчитать вторую половину  
        count_of_work_days_for_first_half  = count_of_work_days #считаем кол-во рабочих дней для первых 15 дней и "запоминаем их"
        count_of_work_days = 0 #обнуляем счетчик, чтобы посчитать вторую половину
        for i in sheet[f'E{r+2}:T{r+2}'][0]:   #банальный подсчет рабочих дней для второй половины месяца
            Workday = i.value
            print(Workday)
            if(Workday != 'ДР' and Workday != 'Я' and Workday != 'В' and  Workday != None and Workday != "О" and Workday != "Б"):
                day_days_first_half+=1
                night_days_first_half+=1      
            elif Workday == 'ДР':
                count_of_work_days += 1
            elif Workday == 'Я':
                print("ALLAH")
                count_of_work_days += 1
            #print("count_of_work_days   ", count_of_work_days)        
            #return  count_of_work_days    
        print("------------------------------")
        #заполняем все нужные клетки
        if (file != "Отдел информационных технологий.xlsx"):
            sheet[cell_function_days_U(cell_num_days+2)] = count_of_work_days 
            
        else:
            sheet[cell_function_days_U(cell_num_days+2)] = str("{:.0f}".format(count_of_work_days + day_days_second_half)) + "/" + str("{:.0f}".format(night_days_second_half))
            
        for i in sheet[f'E{r+3}:T{r+3}'][0]:   #подсчет часов
            if(i.value !=None): #в месяце могут быть не все 31, а значит могут быть пустые клетки, тобиш None, их складывать нельзя
                Workday = i.value
                if(isinstance(Workday,int)==True or isinstance(Workday,float)==True):
                    count_of_work_hours += Workday                                     
                else:
                    night_hours_second_half+=4 * day_fraction_for_lazy_ass
                    day_hours_second_half += 2 * day_fraction_for_lazy_ass
        
        if (file != "Отдел информационных технологий.xlsx"):
            sheet[cell_function_hours_U(cell_num_hours+2)] =  count_of_work_hours
        else:   
            sheet[cell_function_hours_U(cell_num_hours+2)] =  str("{:.0f}".format(count_of_work_hours+day_hours_second_half)) + "/" + str("{:.0f}".format(night_hours_second_half))
        
        if (file != "Отдел информационных технологий.xlsx"):
            sheet[cell_function_hours_V(cell_num_hours +1)] = count_of_work_hours + count_of_work_hours_half
        else:   
             sheet[cell_function_hours_V(cell_num_hours +1)] = str("{:.0f}".format(count_of_work_hours + count_of_work_hours_half + day_hours_first_half + day_hours_second_half)) + "/" + str("{:.0f}".format(night_hours_first_half+night_hours_second_half)) #сумма двух половин месяца
       
        
        if (file != "Отдел информационных технологий.xlsx"):
            sheet[cell_function_days_V(cell_num_days)] = count_of_work_days + count_of_work_days_for_first_half
        else: 
            sheet[cell_function_days_V(cell_num_days)] = str("{:.0f}".format(count_of_work_days + count_of_work_days_for_first_half + day_days_first_half + day_days_second_half)) + "/" + str("{:.0f}".format(night_days_first_half + night_days_second_half))#сумма двух половин месяца
        count_of_work_hours = 0
        count_of_work_days  = 0
        night_days_first_half = 0 
        night_days_second_half = 0
        day_hours_first_half = 0 
        day_hours_second_half = 0
        day_days_first_half = 0 
        day_days_second_half = 0
        night_hours_first_half = 0 
        night_hours_second_half = 0
            
calendar = ProdCalendar(locale='ru') #добавляем производственный календарь с Российскими выходными/праздниками

staff = pd.read_excel('Штат08.10.2021.xls')  # считываем файл со штатом сотрудников

staff['full_name_dep']=(staff["Фамилия"] + " " + staff["Имя"] + " " + staff["Отчество"] + " " 
                        + staff["Штатная должность"]).astype("string") 
#делаем full_name - ФИО + департамент где работает + ставка, на которую работает человек 

DEPARTMENTS = staff['Подразделение (наименование)'].unique() 
#находим все уникальные департменты в столбце "Подразделение(наименование)"
TEMPLATE_SIZE = 100 # ppl per file(кол-во людей, которое может быть в одном файле, если окажется что в департменте работает больше, просто изменить значение)
FIRST_ROW = 18  #первая строчка, откуда начинается заполнение табеля сотрудниками 
#здесь как раз и отлавливаю ошибку, если вдруг окажется что лист очень маленький, а народу в департменте много
MAX_PPL = staff.groupby('Подразделение (наименование)')['full_name_dep'].count().max()
#ppl = people
if TEMPLATE_SIZE < MAX_PPL:
    raise ValueError('Шаблон табеля меньше, чем кол-во людей в каком-то департаменте!\n Для этого нужно взять шаблон(TableForm.xlsx) и добавить туда примерно +50% от текущего (100=>150=>225 etc)')
    
LAST_ROW = FIRST_ROW + TEMPLATE_SIZE*4 - 1  # последняя пустая строчка в  шаблоне табеля
"""
cell height for the name of the person signing the document = CHFTNOTPSTD = CTD
высота ячейки, для названия должности сотрудника, который будет подписывать документ
данную величину можно регулировать по усмотрению(x пикселей)
------------------------------------------------
jump - через если клетка слита через 4 ячейки, т это значит что после функии нужно идти на 4 вниз, к следующей
start_col - первый столбец, откуда все и начинается 
last_col - последний столбец
"""
def unmerge_cells_fun(jump, start_col, last_col):
        row = r
        for row in range(row, LAST_ROW, jump):
            for col in[(start_col,last_col)]:
                sheet.unmerge_cells(start_row=row, start_column=col[0], end_row=row + jump - 1, end_column=col[1])

#все 4 снизу делают почти тоже самое, что и start_col_fun(start_col). Имеет более локальный вид
def cell_function_days_U(cell_num_days): #добавление отработанных дней в колонку U
        days_cell = str("U") + str(cell_num_days)       #создаем номер ячейки, куда будем заполнять данные
        return days_cell
    
def cell_function_hours_U(cell_num_hours):  #добавление отработанных часов в колонку U
        hours_cell = str("U") + str(cell_num_hours)
        return hours_cell
        
def cell_function_days_V(cell_num_days): #добавление отработанных дней в колонку V
        days_cell_2 = str("V") + str(cell_num_days)  #создаем номер ячейки, куда будем заполнять данные
        return days_cell_2
      
def cell_function_hours_V(cell_num_hours): #добавление отработанных часов в колонку V
        hours_cell_2 = str("V") + str(cell_num_hours) 
        return hours_cell_2    
"""    
start_col - первый столбец, откуда все и начинается 
last_col - последний столбец
fpos_row -первая строка, где начинаем соединять 
lpos_row - послдная строка, где заканчиваем соединять 
k - некоторые вещи нам нужно соединить один раз(без цикла), и удобнее просто указать их место
"""        
def merge_cells_fun(start_col, last_col, fpos_row, lpos_row, k):
    row = r + k
    for col in[(start_col,last_col)]:
        sheet.merge_cells(start_row=row + fpos_row, start_column=col[0], end_row=row + lpos_row, end_column=col[1])
        
"""
используется для добавления техта в поля для департмента, дата создания файла, с какой по какой период. Также выравнивание по центру
"""        
def add_into_merged_cell(start_col,number_of_the_row,text):
    top_left_cell = sheet[f'{start_col_fun(start_col)}{number_of_the_row}'] #обозначаем клетку, куда хотим добавить название 
    top_left_cell.value = text #добавляем название
    top_left_cell.alignment = Alignment(horizontal="center", vertical="center")    #выравниваем по цетру        
    if (top_left_cell.value == None):
        raise ValueError('В клетку ничего не было записано! Она пустая!')

"""
превращается нумерацию столбцов в буквы, чтобы можно было использовать в коде
Например : 22=> V ; 27 => AA
"""
def start_col_fun(start_col):
    if(start_col == 0):
        raise ValueError('столбец равен нулю! Такого быть не может, ведь нумерация с единицы/первого. Это означает, что ошибка в системе, где-то start_col обнуляется')
    if(start_col>26): 
        start_col -=26
        start_col_fun_2 = "A" + str(chr(64+start_col))
    else:
        start_col_fun_2 = str(chr(64+start_col))
    return start_col_fun_2

"""
проставляет обычные 5 по 8, с поправкой на ставку. Если выходной (по производственному календарю), то проставляет 0 
"""
def standart_5_8_day_fraction(day_in_the_month,column_for_work_days, day_fraction_for_lazy_ass, k):
    if calendar.check(datee(2021, 11, day_in_the_month))== DayType.WORKING:
                        sheet.cell(r+k,column_for_work_days).value = "Я"
                        sheet.cell(r+k+1,column_for_work_days).value = 8 * day_fraction_for_lazy_ass
                        if (sheet.cell(r+k,column_for_work_days).value == "Я" and sheet.cell(r+k,column_for_work_days).value == 0):
                            raise ValueError('Данные день отмечен как рабочий, при этом нет рабочих часов. Скорее всего сбит days_count или row_shift. Возпользуйтесь дебагом. Для это вбейте в поиск debug и и запустите какой-нибудь, благо их много')
    else:
                        sheet.cell(r+k,column_for_work_days).value = "В"
                        sheet.cell(r+k+1,column_for_work_days).value = 0
                        if (sheet.cell(r+k,column_for_work_days).value == "В" and sheet.cell(r+k+1,column_for_work_days).value != 0):
                                raise ValueError('Хоть и стоит выходной, в данный день почему-то имеются рабочие часы. Либо опечатка строчкой выше, либо сбит days_count(может что-то еще)')
#ищем два выходных подряд, чтобы туда поставить ночную смену, если она была 15 или 31 числа 
def LF_free_space_for_night_shift(column_for_work_days,day_fraction_for_lazy_ass, k):
    for i in range(5, 20):
        if(sheet.cell(r+k,column_for_work_days).value == "В" and sheet.cell(r+k,column_for_work_days+1).value == "В"):
                sheet.cell(r+k,column_for_work_days).value = "Я/Н"
                sheet.cell(r+1+k,column_for_work_days).value = str("{:.0f}".format(2*day_fraction_for_lazy_ass)) + "/" +  str("{:.0f}".format(2*day_fraction_for_lazy_ass))
                sheet.cell(r+k,column_for_work_days+1).value = "Я/Н"
                sheet.cell(r+1+k,column_for_work_days+1).value = str("{:.0f}".format(2*day_fraction_for_lazy_ass)) + "/" + str("{:.0f}".format(6*day_fraction_for_lazy_ass))
                break
        column_for_work_days+=1
                            
'''
все что внизу - наработчки, которые пока не используются
соединяю поля, куда будут вбиваться подписи, и заоодно сраззу эти подписи
пока что не работает
'''
def last_col_fun(last_col):
    if(start_col == 0):
        raise ValueError('столбец равен нулю! Такого быть не может, ведь нумерация с единицы/первого. Это означает, что ошибка в системе, где-то last_col обнуляется')
    last_col_fun_2 = str(chr(64+last_col))
    return last_col_fun_2
       
def merge_cells_fun_inside(start_col, last_col, fpos_row, lpos_row, k,text):
    row = r + k   
    for col in[(start_col,last_col)]:
        sheet.merge_cells(start_row=row + fpos_row, start_column=col[0], end_row=row + lpos_row, end_column=col[1]) 
    
    add_into_merged_cell(start_col,r+k,text)   
for DEPARTMENT in DEPARTMENTS:    
    #создаем файл для каждого департмента 
    file = f'{DEPARTMENT}.xlsx'
    copy('TableForm.xlsx', file) #название можно поменять
    wb = xl.load_workbook(file,data_only=True)
    sheet = wb['Табель'] #название листа, можно поменять на нужный в шаблоне на входе (то есть в TableForm.xlsx)
    
    add_into_merged_cell(2,5,DEPARTMENT) #добавляем название департмента 
    """
    add_into_merged_cell(24,9,calendar.previous(date.today(), DayType.WORKING))#добавляем дату составления
    add_into_merged_cell(27,9,calendar.previous(date(2021, 11, 1), DayType.WORKING))#добавляем с какого начинается отчет
    add_into_merged_cell(29,9,calendar.previous(date(2021, 11, 30), DayType.WORKING))#добавляем по какое число отчет
    """
    # добавляем всех сотрудников в файл
    people = staff.loc[staff["Подразделение (наименование)"] == DEPARTMENT,                # фильтр по отделу
                                            ['full_name_dep', 'Табельный номер','Ставка (на текущую дату)']] # столбцы которые нам нужны
    people = people.sort_values('full_name_dep')
    n = people.shape[0] # n = кол-во людей 
    r = FIRST_ROW       # r = текущая строчка    
    cell_num_days = 18           #номер ячейки, сколько дней сотрудник появлялся на работе или работал дистанционно 
    cell_num_hours = 19        #номер ячейки, куда мы будем добавлять сумму всех отработанных часов
    numbering = 1                   #нумерация людей в списке 
    
    count_of_work_days = 0   #счетчкий, отвечающий за подсчет рабочих дней у сотруднкика 
    count_of_work_hours = 0 #счетчкий, отвечающий за подсчет рабочих часов у сотруднкика 
    
    row_shift = 2 #используется для файла IT_смен; отвечает за строчки людей, то есть первая строка наименования столбоцв, дальше люди по списку
    
    night_days_first_half = 0 #считает кол-во дней, в которые работали ночбю (одна ночная смена = 2 ночных дня)
    night_days_second_half = 0
    day_days_first_half = 0 #1. ужасное название ; 2. считает кол-во дней, в которые работали днем (также для ночной смены 1 к 2)
    day_days_second_half = 0
    night_hours_first_half = 0 #переменная, отвечающая за ночные часы (их нужно считать отдельно, т.к. оплачиваются большим кол-во $)
    night_hours_second_half = 0
    day_hours_first_half = 0 #ночная смена, днем оплачивается также, нужно просто для их подсчета
    day_hours_second_half = 0
      
    for person in people.itertuples():
        sheet.cell(r, 3).value = person[1]   # ФИО + должность 
        sheet.cell(r, 4).value = person[2]   # Табельный номер
        sheet.cell(r, 1).value = person[3]  # добавление ставки, которую имеет сотрудник
        day_fraction_for_lazy_ass = person[3] # добавляем ставку, чтобы потом на нее умножить часы (то есть при стандартных 8ч и 0.5 ставке будет 8 * 0.5 = 4)
        sheet.cell(r, 2).value = numbering   #нумерация сотрудников в файле
        
        if(day_fraction_for_lazy_ass == 0):
            print(person[1])
            raise ValueError("Ставка равна нулю! что-то не так в таблице всех сотрудников")
            
        
        numbering += 1
        
        day_in_the_month = 1 #счетчик дня в месяце
        amount_days_in_the_month = 30
        #кол-во дней в месяце. Пока не придумал как автомотизировать, чтобы прога сама выбирала кол-во дней в зависимости от месяца
        
        column_for_work_days = 5 #столбец, откуда мы начинаем проставлять рабочие/выходные дни 
        
        days_count=1 #идет счет дней, чтобы после 15-ого дня перейти на новую строчку для второй половины месяца
        #если не IT_файл. То есть берем департменты где надо просто проставить 5 по 8
        #if (file != "Отдел информационных технологий.xlsx"):
        for i in range(1, amount_days_in_the_month): #весь месяц, от первого до последнего дня  
            while(days_count<30+1): #вторая половина 
                while(days_count < 15+1): #первая половина
                    standart_5_8_day_fraction(day_in_the_month,column_for_work_days, day_fraction_for_lazy_ass, 0) #проставление рабочих дней и часов
                    day_in_the_month+=1
                    days_count+=1
                    column_for_work_days+=1         
                if(days_count == 16): column_for_work_days = 5 #начинаем с 5-ого столбца, идем до 22; сделано так, потому что месяц в таблице поделен на 2 части, и приходится начинать с "начала"
                standart_5_8_day_fraction(day_in_the_month,column_for_work_days, day_fraction_for_lazy_ass, 2)   
                day_in_the_month+=1
                days_count+=1
                column_for_work_days+=1         
                
                
        else:   
            #здесь разбирается случай для IT_отдела, ведь у него не 5 по 8, а проставленные смены
            work_shift = pd.read_excel('ОИТ.xlsx')
            file_IT = 'ОИТ.xlsx'
            wb_IT = xl.load_workbook(file_IT,data_only=True)
            sheet_IT = wb_IT['2021 сентябрь']
            #подгружаем excel файл со сменами
            first_shift_day_col = 8 #сначала идет инфа о сотрудниках, сама же инфа о сменах начинается с 8-ого столбцв
            shitf_day = 1 #shift - смена, будет встречаться часто; что же касается переменной, отвечает за счетчик дней внутри  файла со сменами
             
                
            
            for i in range(1, amount_days_in_the_month): #весь месяц, от первого до последнего дня         
                while(days_count<30+1):
                    while(days_count < 15+1): #первая половина                                          
                        #если стоит : отпуск, выходной, ничего не стоит (None), то ставим 0 часов и выходной(можно и разбить чтобы отпуска в финальном табеле ставились)
                        if(sheet_IT[f'{start_col_fun(first_shift_day_col)}{row_shift}'].value == "О"): 
                            sheet.cell(r,column_for_work_days).value = "О"
                            sheet.cell(r+1,column_for_work_days).value = 0  
                        elif(sheet_IT[f'{start_col_fun(first_shift_day_col)}{row_shift}'].value == None):
                            standart_5_8_day_fraction(day_in_the_month,column_for_work_days, day_fraction_for_lazy_ass, 0)
                            
                        elif (sheet_IT[f'{start_col_fun(first_shift_day_col)}{row_shift}'].value == "В"):                          
                            sheet.cell(r,column_for_work_days).value = "В"
                            sheet.cell(r+1,column_for_work_days).value = 0  
                            if (sheet.cell(r,column_for_work_days).value == "В" and sheet.cell(r+1,column_for_work_days).value != 0):
                                raise ValueError('Хоть и стоит выходной, в данный день почему-то имеются рабочие часы. Либо опечатка строчкой выше, либо сбит days_count(может что-то еще)')
                        
                        #3, или же обычная 8 часовая смена
                        elif int(sheet_IT[f'{start_col_fun(first_shift_day_col)}{row_shift}'].value) == 3:
                            
                            sheet.cell(r,column_for_work_days).value = "Я"
                            sheet.cell(r+1,column_for_work_days).value = 8 * day_fraction_for_lazy_ass
                            if (sheet.cell(r,column_for_work_days).value == "Я" and sheet.cell(r+1,column_for_work_days).value == 0):
                                raise ValueError('Данные день отмечен как рабочий, при этом нет рабочих часов. Скорее всего сбит days_count или row_shift. Возпользуйтесь дебагом. Для это вбейте в поиск debug и и запустите какой-нибудь, благо их много')
                                               
                        #2, или же ночная смена. Тут нужно код доделать, пока просто поставил 20 часов. Там все сложнее, жду уточнений что делать
                        elif int(sheet_IT[f'{start_col_fun(first_shift_day_col)}{row_shift}'].value) == 2:
                            
               
                            #бывает такое, что в 15 день у сотрудника ночная смена. В таком случае нужно поставить 2/6 не в след клетку, а перенести на след половину месяца
                            if(days_count !=15):
                                sheet.cell(r,column_for_work_days).value = "Я/Н"
                                sheet.cell(r+1,column_for_work_days).value =  str("{:.0f}".format(2*day_fraction_for_lazy_ass)) + "/" +  str("{:.0f}".format(2*day_fraction_for_lazy_ass))
                            
                                day_in_the_month+=1 
                                column_for_work_days+=1
                                days_count+=1
                                shitf_day +=1
                                first_shift_day_col +=1
                                
                                sheet.cell(r,column_for_work_days).value = "Я/Н"
                                sheet.cell(r+1,column_for_work_days).value =  str("{:.0f}".format(2*day_fraction_for_lazy_ass)) + "/" +  str("{:.0f}".format(6*day_fraction_for_lazy_ass))
                            else:
                                sheet.cell(r,column_for_work_days).value = "В"
                                sheet.cell(r+1,column_for_work_days).value = 0
                                
                                column_for_work_days = 5
                                LF_free_space_for_night_shift(column_for_work_days,day_fraction_for_lazy_ass, 0)
                                
                                                                                     
                        #1, или же обычная(вернется живым не каждый...) 12 часовая смена
                        elif(int(sheet_IT[f'{start_col_fun(first_shift_day_col)}{row_shift}'].value)) == 1:                           
                                sheet.cell(r,column_for_work_days).value = "Я"
                                sheet.cell(r+1,column_for_work_days).value = 12 * day_fraction_for_lazy_ass
                                
                   
                        day_in_the_month+=1 
                        column_for_work_days+=1
                        days_count+=1
                        shitf_day +=1
                        first_shift_day_col +=1
                        
                                     
                    if(days_count == 16): column_for_work_days = 5 #начинаем с 5-ого столбца, идем до 22
                    """
                    print(day_in_the_month)   
                    print(column_for_work_days)   
                    print(days_count)   
                    print(shitf_day)   
                    print(first_shift_day_col)   
                    print("-------------------------------")
                    """
                        
                     #если стоит : отпуск, выходной, ничего не стоит (None), то ставим 0 часов и выходной(можно и разбить чтобы отпуска в финальном табеле ставились)
                    if(sheet_IT[f'{start_col_fun(first_shift_day_col)}{row_shift}'].value == "О"):
                        print("lol")
                    elif(sheet_IT[f'{start_col_fun(first_shift_day_col)}{row_shift}'].value == None):
                            standart_5_8_day_fraction(day_in_the_month,column_for_work_days, day_fraction_for_lazy_ass, 2)
                            
                    elif (sheet_IT[f'{start_col_fun(first_shift_day_col)}{row_shift}'].value == "В"):                           
                            sheet.cell(r,column_for_work_days).value = "В"
                            sheet.cell(r+1,column_for_work_days).value = 0  
                            if (sheet.cell(r,column_for_work_days).value == "В" and sheet.cell(r+1,column_for_work_days).value != 0):
                                raise ValueError('Хоть и стоит выходной, в данный день почему-то имеются рабочие часы. Либо опечатка строчкой выше, либо сбит days_count(может что-то еще)')
                        
                        #3, или же обычная 8 часовая смена
                    elif int(sheet_IT[f'{start_col_fun(first_shift_day_col)}{row_shift}'].value) == 3:
                            
                            sheet.cell(r,column_for_work_days).value = "Я"
                            sheet.cell(r+1,column_for_work_days).value = 8 * day_fraction_for_lazy_ass
                            if (sheet.cell(r,column_for_work_days).value == "Я" and sheet.cell(r+1,column_for_work_days).value == 0):
                                raise ValueError('Данные день отмечен как рабочий, при этом нет рабочих часов. Скорее всего сбит days_count или row_shift. Возпользуйтесь дебагом. Для это вбейте в поиск debug и и запустите какой-нибудь, благо их много')
                                               
                        #2, или же ночная смена. Тут нужно код доделать, пока просто поставил 20 часов. Там все сложнее, жду уточнений что делать
                    elif int(sheet_IT[f'{start_col_fun(first_shift_day_col)}{row_shift}'].value) == 2:
                            print((sheet_IT[f'{start_col_fun(first_shift_day_col)}{row_shift}']))
                            print(days_count)       
               
                            #бывает такое, что в 15 день у сотрудника ночная смена. В таком случае нужно поставить 2/6 не в след клетку, а перенести на след половину месяца
                            if(days_count != amount_days_in_the_month):
                                sheet.cell(r,column_for_work_days).value = "Я/Н"
                                sheet.cell(r+1,column_for_work_days).value =  str("{:.0f}".format(2*day_fraction_for_lazy_ass)) + "/" +  str("{:.0f}".format(2*day_fraction_for_lazy_ass))
                            
                                day_in_the_month+=1 
                                column_for_work_days+=1
                                days_count+=1
                                shitf_day +=1
                                first_shift_day_col +=1
                                
                                sheet.cell(r,column_for_work_days).value = "Я/Н"
                                sheet.cell(r+1,column_for_work_days).value =  str("{:.0f}".format(2*day_fraction_for_lazy_ass)) + "/" +  str("{:.0f}".format(6*day_fraction_for_lazy_ass))
                            else:
                                sheet.cell(r,column_for_work_days).value = "В"
                                sheet.cell(r+1,column_for_work_days).value = 0
                                
                                column_for_work_days = 5
                                LF_free_space_for_night_shift(column_for_work_days,day_fraction_for_lazy_ass, 2)
                                                                                     
                        #1, или же обычная(вернется живым не каждый...) 12 часовая смена
                    elif(int(sheet_IT[f'{start_col_fun(first_shift_day_col)}{row_shift}'].value)) == 1:                           
                                sheet.cell(r,column_for_work_days).value = "Я"
                                sheet.cell(r+1,column_for_work_days).value = 12 * day_fraction_for_lazy_ass
                                   
                    day_in_the_month+=1 
                    column_for_work_days+=1
                    days_count+=1
                    shitf_day +=1
                    first_shift_day_col +=1
                    #row_shift+=1
                                         
                                
        #обнуляем счетчики, чтобы потом применить их на след человеке
        row_shift +=1
        column_for_work_days = 5
        day_in_the_month = 1 
        days_count = 1
        column_for_work_days = 1 
        shitf_day = 1
        first_shift_day_col = 8    
        
       #берем 'r' как первую строчку 

        #sutuation_obosration(night_days_second_half,count_of_work_days,day_days_first_half,night_days_first_half,count_of_work_hours,day_hours_first_half,night_hours_first_half,day_hours_second_half,night_hours_second_half,day_days_second_half,r)
        count_of_work_hours = 0
        count_of_work_days  = 0
        
        #переход на следующею ячейку, то есть на следуещего сотрудника по списку
        
        
        night_days_first_half = 0 
        night_days_second_half = 0
        day_hours_first_half = 0 
        day_hours_second_half = 0
        day_days_first_half = 0 
        day_days_second_half = 0
        night_hours_first_half = 0 
        night_hours_second_half = 0
        
        r += 4 #добавляем через 4 клетки нового сотрудника 
    row = r #возвращеем row в изначальное состояние, чтобы сделать тотже цикл с другими объектами    
    """
    В кратце зачем это нужно :
    НЕЛЬЗЯ(пока я конечно не воспользовался Aligment, но он не всегда работает) заполнять слитые ячейки в Excel
    Точнее можно, но только с помощью VBA, не Python
    для это я разъединяю все ячейки, потом записываю в них что мне нужно и сливаю обратно
    """
    unmerge_cells_fun(2,22,22)            
    unmerge_cells_fun(4,2,2)
    unmerge_cells_fun(4,3,3)
    unmerge_cells_fun(4,4,4)
       
    merge_cells_fun(22,22,0,0,0)  
    merge_cells_fun(2,3,1,1,0)
    merge_cells_fun(4,5,1,1,0)
    #merge_cells_fun_inside(4,5,1,1,1,text) #попытка релиозавать кое-что. Описано в самом низу в функциях
    merge_cells_fun(7,9,1,1,0)
    merge_cells_fun(11,19,1,1,0)
    merge_cells_fun(23,24,1,1,0)
    merge_cells_fun(28,29,1,1,0)
    merge_cells_fun(23,24,4,4,0)
    merge_cells_fun(28,29,4,4,0)  
    # все это сливает ячейки под таблицей, где подписи и прочее
    merge_cells_fun(4,5,0,0,2)
    merge_cells_fun(7,9,0,0,2)
    merge_cells_fun(11,19,0,0,2)
    merge_cells_fun(23,24,0,0,5)
    merge_cells_fun(23,24,0,0,2)
    merge_cells_fun(28,29,0,0,2) 
    merge_cells_fun(28,29,0,0,5)
    merge_cells_fun(21,22,1,1,0)
    merge_cells_fun(21,22,1,1,3)
       
    CTD = 33 #в функциях в самом начале написано что это, если коротко - кол-во пикселей (ширина/высота) строчки в excel 
    rd = sheet.row_dimensions[r+1]   #расширяем ячейку для "Ответственное лицо" и "Руководитель структурного подразделения"
    rd.height = CTD              
    rd1 = sheet.row_dimensions[r+4]  #расширяем ячейку для "Работник кадровой службы"
    rd1.height = CTD            
     
    sheet.delete_rows(r, LAST_ROW + 1 - r) # удаляем неиспользованные строки 
    #корректируем print area чтобы она соотвестовала финальному результату, который нам нужен (без пустых пропусков и т.п)
    sheet.print_area = sheet.calculate_dimension()
    wb.save(file)   #сохраняем все изменения   
    # нужен для теста (сделает один файл), удалите break если надо чтобы сделало все департаменты (весь цикл)
    
   
    #переход на следующею ячейку, то есть на следуещего сотрудника по списку
        

    wb.save(file)
    print(file, "был успешно создан!")
    
    break


dict_tabel_spec = {
    'vacation' : ('Отпуск с', 'Отпуск по', 'отпуск', 'Вид отпуска',
     'Обозначения в файле  "Отпуск"  столбец "Вид отпуска"'),
    'seak_leave' : ('Дата начала', 'Дата окончания', 'больничный', 'Вид пособия',
     'Обозначения в файле "Больничный" столбец "Вид пособия"'),
    'remote_work' : ('С', 'По' , 'дистанционная работа')
}


# TABEL_FILENAME = 'Дирекция (!).xlsx'
"""TABEL_FILENAME_LIST = [
                        'Дирекция (!).xlsx',
                        'Юридический отдел (!).xlsx',
                        'Отдел рентгенорадиологического технического контроля 12.10 (1).xlsx',
                        'Отдел информационных технологий.xlsx',
                        'Сектор обучения и развития персонала (!).xlsx',
                        'Планово-экономический отдел (!).xlsx',
                        'Секретариат (!).xlsx',
                        'Центр по внедрению медицинских технологий (!).xlsx',
                        'Отдел кадров (!).xlsx',
                        'Отдел закупок (!).xlsx'
                        ]"""

DATASOURCE_FILENAME_1 = 'отпуска.xls'
TABEL_FILENAME = 'ОИТ Табель до 15.10.xlsx'
#TABEL_FILENAME_LIST = ['ОИТ Табель до 15.10.xlsx']
TABEL_FILENAME_LIST = ['Отдел информационных технологий.xlsx']
DATASOURCE_FILENAME_2= '1. Больничный.xlsx'
#DATASOURCE_FILENAME = '3. ДР.xlsx'
CODES_FILENAME = '5. Условные обозначения.xlsx'


fda = ('id', *[str(i) for i in range(1,16)])
sda = ('16', *[str(i) for i in range(17,32)])
ind = ('1','2')
date = datetime.datetime.today().strftime('%d.%m.%Y')





# Получение даты из табеля
def get_date_value():
    wb = load_workbook(filename = TABEL_FILENAME, read_only = True, data_only = True)
    ws = wb['Табель']
    tad = ws['AA9'].value
    #dt = tad.date().month
    wb.close()
    
    return tad


DAT = get_date_value()


def dt(day, month, year): 
    return pd.Timestamp(day = int(day), month = month, year = year)


def date_func(day, month, year):
    return dt(day, month, year) if (day.isdigit()) else dt(1, 1, 1991)


def cell_level(day):
    return day // 16


def last_day_digit(ind):
    return 30 if ind == 31 else ind


def line_num(count, ind):
    p = last_day_digit(ind)
    return 18 + 4 * count + 2 * cell_level(p)


def col_num(ind):
    p = ind % 16 + cell_level(ind)
    return 4 + p


def fill_cell(ws, count, index, value):
    line = line_num(count[0], int(index))
    col = col_num(int(index))
    cell = ws.cell(column = col, row = line)
    cell.value = value
    cell = ws.cell(column = col, row = line + 1)
    cell.value = 0
    return None


# Проверка условий даты в соответствии с названием столбца даты ("от", "до") в файле-источнике
def is_fromto_date(df, tabel_spec, fromto, ind, dt = DAT):
    date_rep = pd.to_datetime(dt, format='%d.%m.%Y')
    month, year = date_rep.month, date_rep.year
    if fromto == tabel_spec[0]:
        return df[fromto] <= date_func(ind, month, year)
    elif fromto == tabel_spec[1]:
        return df[fromto] >= date_func(ind, month, year)


def match_dates(df_tbl, ind, df_sl, tabel_spec, type, df_codes):
    df = df_sl.loc[df_sl['Табельный номер'] == df_tbl['id']].loc[
        is_fromto_date(df_sl, tabel_spec, tabel_spec[0], ind)].loc[
        is_fromto_date(df_sl, tabel_spec, tabel_spec[1], ind)]
    if(not df.empty):
        df1 = df_codes.loc[lambda x: x[tabel_spec[4]] == df[tabel_spec[3]][df[tabel_spec[3]].index[0]]]
        print(df1['Буквенный код'][df1['Буквенный код'].index[0]])
        return df1['Буквенный код'][df1['Буквенный код'].index[0]]
    return None


def get_matching(matching_check_func, df_tbl, ind, df_sl, tabel_spec, type, df_codes):
    if(ind == 'id'):
        return None
    return matching_check_func(df_tbl, ind, df_sl, tabel_spec, type, df_codes)


def get_offhours(df_tbl, index, df_sl, tabel_spec, count, ws, type, df_codes):

    for ind in df_tbl.index:
        value = get_matching(match_dates, df_tbl, ind, df_sl, tabel_spec, type, df_codes)
        if(value != None):
            fill_cell(ws, count, ind, value)


def offhours(df_tbl, df_sl, i, tabel_spec, ws, type, df_codes):
    if(str(df_tbl[0]) == 'nan'):
        return
    get_offhours(
        df_tbl, ind, df_sl, tabel_spec, i, ws, type, df_codes
        )

    i[0] += 1


def create_df(df, loc_a, loc_b, columns, index=ind, i=0):
    new_df = df.iloc[i+loc_a[0]:i+loc_a[1], loc_b[0]:loc_b[1]]
    new_df.columns = columns
    new_df.index = index
    return new_df




def main():
    for tabel_filename in TABEL_FILENAME_LIST:
        print(f'\nОбрабатывается {tabel_filename}')
        df = pd.read_excel (tabel_filename, "Табель", usecols="D:T", na_values=['NaN'], skiprows=16)
        df_sl1 = pd.read_excel(DATASOURCE_FILENAME_1)
        df_sl2 = pd.read_excel(DATASOURCE_FILENAME_2)
        df_codes = pd.read_excel(CODES_FILENAME)

        df1 = create_df(df, (0,2), (0, 16), fda)
        df2 = create_df(df, (2,4), (1, 17), sda)

        for i in range(4, len(df.index), 4):
            if(len(df.index) - i < 4 or str(df.iloc[i, 0]) == 'nan'):
                break
            ind = [(i / 2) + 1, (i / 2) + 2]
            df11 = create_df(df, (0,2), (0, 16), fda, ind, i)
            df1 = pd.concat([df1,df11], axis = 0)
            df22 = create_df(df, (2,4), (1, 17), sda, ind, i)
            df2 = pd.concat([df2,df22], axis = 0) 

        df3 = pd.concat([df1,df2], axis = 1)

        if(not df3.empty):
            wb = load_workbook(filename = tabel_filename)
            ws = wb['Табель']
            count = [0]
            df3.apply(offhours, axis = 1, args = (df_sl1, count, dict_tabel_spec['vacation'], ws, 'vacation', df_codes))
            df3.apply(offhours, axis = 1, args = (df_sl2, count, dict_tabel_spec['seak_leave'], ws, 'seak_leave', df_codes))
            wb.save(f"{tabel_filename}_output_{date}.xlsx")
            wb.close()

main()
