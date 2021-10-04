import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def date(year, month, day):
    d = pd.Timestamp(day = int(day), month = month, year = year)
    return d


def check_seak_leave(df_sl, ws_timesheet):
    i = 0
    for row in ws_timesheet.iter_rows(min_row=19, values_only=True):
        for ind in range (4, len(row) - 1):
            if(not df_sl.loc[lambda x: x['Табельный номер'] == row[3]].empty):
                if(not df_sl.loc[lambda x: 
                x['Дата окончания'] <= date(2021, 8, row[ind])].loc[lambda x:    #how can I define year and date?
                x['Дата окончания'] >= date(2021, 8, row[ind])].empty):
                    d = ws_timesheet.cell(row = i + 2,column = ind)
                    d.value = 0
                    d.fill = PatternFill(start_color = "70AD47", end_color = "70AD47", fill_type = 'solid')
                    d = ws_timesheet.cell(row = i + 2,column = ind)
                    d.value = "Б"
        i+=1
    return ws_timesheet


df_sl = pd.read_excel("1. Больничный.xlsx")
wb2 = load_workbook('6. Форма табеля.xlsx')
ws1 = wb2["0504421"]



ws1 = check_seak_leave(df_sl, ws1)


wb2.save("New.xlsx")
